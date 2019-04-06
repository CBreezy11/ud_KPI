import openpyxl



class ReadXlsx():
      
    def __inti__(self):
        pass
     
    def get_sheet(self):
        filename = input("Enter the filename: ")
        wb = openpyxl.load_workbook('{}'.format(filename+".xlsx"))
        ws = wb.active
        return ws, filename

                            
    def get_servers(self, ws):
        server1 = []
        servers = []
        for row in range(2, ws.max_row):
            item = ws['B' + str(row)].value
            server1.append(item)
        for y in server1:
            if y is not None:
                if "Server:" in str(y):
                    servers.append(y)
        return servers

    def load_ranges(self, ws, servers):
            serverRows = []
            for cell in ws['B']:
                    if cell.value is not None:
                            for i in servers:
                                    if i == cell.value:
                                            serverRows.append(cell.row)
            serverRows.append(ws.max_row)
            return serverRows
                    
    def load_data(self, d, ws, serverRows):
        for row in range(serverRows[0], serverRows[1]):
            item = ws['C' + str(row)].value
            quant = ws['D' + str(row)].value
            d[item] = quant

    def make_dict(self, ws, servers, serverRows):
            master_list = []
            for server in servers:
                    server = {server: {}}
                    self.load_data(server, ws, serverRows)
                    master_list.append(server)
                    del serverRows[0]
            return master_list
                    
    def get_32(self, master_list):
            i = 0
            while i < len(master_list):
                    master_list[i]['32'] = 0
                    master_list[i]['16'] = 0
                    master_list[i]['vodka'] = 0
                    master_list[i]['VP'] = 0
                    master_list[i]['well'] = 0
                    for y in master_list[i]:
                            if y is not None:
                                    if "_32" in y:
                                            master_list[i]['32'] += master_list[i][y]
                                    if "16" in y:
                                            master_list[i]['16'] += master_list[i][y]
                                    #elif "LEVEL UP" in y:
                                            #sumCocktail += master_list[i][y]
                                    #elif "SQUAD UP" in y:
                                            #sumCocktail += master_list[i][y]
                                    if "tito" in str(y).lower():
                                            master_list[i]['VP'] += master_list[i][y]
                                            #master_list[i]['vodka'] += master_list[i][y]
                                    if "goose" in str(y).lower():
                                            master_list[i]['VP'] += master_list[i][y]
                                            #master_list[i]['vodka'] += master_list[i][y]
                                    if "sky" in str(y).lower():
                                            master_list[i]['well'] += master_list[i][y]
                                            #master_list[i]['vodka'] += master_list[i][y]
                                    if "Vodka" in str(y).lower():
                                            master_list[i]['vodka'] += master_list[i][y]
                    i += 1

def tryMath(i, n):
    try:
        x32 = round(xmaster[i]['32'] / (xmaster[i]['32'] + xmaster[i]['16']) * 100, 2)
    except:
        x32 = 0
    try:
        y32 = round(ymaster[n]['32'] / (ymaster[n]['32'] + ymaster[n]['16']) * 100, 2)
    except:
        y32 = 0
    try:
        xVP = round(xmaster[i]['VP'] / xmaster[i]['vodka'] *100, 2)
    except:
        xVP = 0
    try:
        yVP = round(ymaster[n]['VP'] / ymaster[n]['vodka'] *100, 2)
    except:
        yVP = 0
    try:
        xw = round(xmaster[i]['well'] / xmaster[i]['vodka'] * 100, 2)
    except:
        xw = 0
    try:
        yw = round(ymaster[n]['well'] / ymaster[n]['vodka'] * 100, 2)
    except:
        yw = 0
    return x32, y32, xVP, yVP, xw, yw
    
def results():
    f1.write("\n\n")
    i = 0
    list = ['32', '16', 'vodka', 'VP', 'well']
    while i < len(xmaster):
        f1.write(xservers[i]+"\n")
        f1.write("{}{}\n".format(f, f2.rjust(50, " ")))
        f1.write(7*"*" + (7*"*").rjust(47, " ")+"\n")
        if xservers[i] in yservers:
            n = yservers.index(xservers[i])
            x, y, z, aa, bb, cc = tryMath(i, n)
            f1.write("32oz beer upsell %: {}".format(x)+"32oz beer upsell %: {}\n".format(y).rjust(35, " "))
            f1.write("Vodka Premium %: {}".format(z)+"Vodka Premium %: {}\n".format(aa).rjust(30, " "))
            f1.write("Vodka Well %: {}".format(bb)+("Vodka Well %: {}\n".format(cc).rjust(37, " ")))
        else:
            f1.write("32oz beer upsell %: {}\n".format(x))
            f1.write("Vodka Premium %: {}\n".format(z))
            f1.write("Vodka Well %: {}\n".format(bb))
        f1.write("\n")
        i += 1


               
x = ReadXlsx()
sheet, f = x.get_sheet()
xservers = x.get_servers(sheet)
xranges = x.load_ranges(sheet, xservers)
xmaster = x.make_dict(sheet, xservers, xranges)
x.get_32(xmaster)

y = ReadXlsx()
sheet2, f2 = y.get_sheet()
yservers = y.get_servers(sheet2)
yranges = y.load_ranges(sheet2, yservers)
ymaster = y.make_dict(sheet2, yservers, yranges)
y.get_32(ymaster)

f1 = open("{}_compareto_{}.txt".format(f, f2), "w+")

results()
f1.close()

