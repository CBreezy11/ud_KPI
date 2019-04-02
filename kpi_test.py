import openpyxl

wb = openpyxl.load_workbook('3.13.xlsx')
ws = wb.active

servers = []

master_list =[]

serverRows = []

def get_servers():
    server1 = []
    for row in range(2, ws.max_row):
        item = ws['B' + str(row)].value
        server1.append(item)
    for y in server1:
        if y is not None:
            if "Server:" in str(y):
                servers.append(y)

def load_ranges():
        for cell in ws['B']:
                if cell.value is not None:
                        for i in servers:
                                if i == cell.value:
                                        serverRows.append(cell.row)
        serverRows.append(ws.max_row)
                
def load_data(d):
    for row in range(serverRows[0], serverRows[1]):
        item = ws['C' + str(row)].value
        quant = ws['D' + str(row)].value
        d[item] = quant

def make_dict():
        for server in servers:
                server = {server: {}}
                load_data(server)
                master_list.append(server)
                del serverRows[0]
                
def get_32():
        i = 0
        while i < len(master_list):
                sum16 = 0
                sum32 = 0
                for y in master_list[i]:
                        if y is not None:
                                if "32" in y:
                                        sum32 += master_list[i][y]
                                elif "16" in y:
                                        sum16 += master_list[i][y]
                results(sum32, sum16, i)
                i += 1

def results(x, y, i):
        print("{} sold {} 32oz beers which is {}%".format(servers[i], x, round((x / (x + y)* 100), 2)))

get_servers()       
load_ranges()
make_dict()
get_32()

                  



  



